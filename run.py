import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from operation import Operations
from bilan_operation import bilan_operations

border_default=Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def val(sheet,x, y):
    return sheet.cell(row=x, column=y).value

# fonction de lecture de fichier pour transformer en liste
def ExcelfileToList(name_file):
    file = name_file + ".xlsx"
    dataframe = openpyxl.load_workbook(file)
    
    # Define variables
    dataframe1 = dataframe.active
    list = []
    cpt=0
    temp_amount=0
    temp_name=""
    temp_no_account =0
    temp_type=0 # 0 = revenue, 1= depense
    string_list = ['Revenus :'] # filtre indésirables
    string_to_switch="Frais d'exploitation :" #mot pour switch entre revenues et dépsenses
    string_to_stop ="Total des frais"
    is_over = 0 #bool
    
    # Iteration dans les row
    for row in range(4, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, 3): # prendre que les 2 premieres col des row
            if(col[row].value != string_to_stop and is_over == 0): 
                if(col[row].value not in string_list and col[row].value != None): #verifier si on a pas de data indésirable (filtre)
                    if(cpt==0):
                        temp_no_account = col[row].value
                        cpt = cpt+1
                    elif(cpt==1):
                            if(col[row].value == string_to_switch):    #switch entre les revenues et les dépenses
                                temp_type = 1
                            temp_name = col[row].value
                            cpt = cpt+1
                    else:
                        temp_amount = col[row].value
                        cpt =0
                        list.append(Operations(temp_no_account,temp_name,temp_type,temp_amount,name_file)) #ajout a la liste
            else:
                is_over=1
    return list

#regroupement des listes
def ReadAllExcel():
    arvida_lst  = ExcelfileToList("SAINT-DOMINIQUE")
    kenogami_lst  = ExcelfileToList("SAINTE-FAMILLE")
    final_lst = arvida_lst + kenogami_lst

    for x in final_lst:
        print(x.name+" - montant : "+str(x.amount))
    return final_lst

#on passe la list a trier dans le excel, la feuille Excel et l'index de commencement 
def WriteRevenus(list, sheet):

    #revenus
    f5 = sheet.cell(row=5, column=6)
    f5.value = "REVENUS"
    f5.font = Font(bold=True,size=14)

     #paroissiens
    f5 = sheet.cell(row=6, column=6)
    f5.value = "  PAROISSIENS"
    f5.font = Font(bold=True,size=12)

    for x in range(0,11):
        tempAccount = sheet.cell(row=6+(x+1), column=6)
        tempAccount.value = list[x].account
        tempAccount.font = Font(bold=True)
        tempAccount.alignment = Alignment(horizontal="center")
        tempAccount.border = border_default

        temp_name = sheet.cell(row=6+(x+1), column=7)
        temp_name.value = list[x].name
        temp_name.border = border_default

        temp_st_do = sheet.cell(row=6+(x+1), column=9)
        temp_st_do.value = list[x].st_dominique_amount
        temp_st_do.border = border_default
        temp_st_do.number_format = '#,##0.00$' 

        temp_st_fam = sheet.cell(row=6+(x+1), column=10)
        temp_st_fam.value = list[x].sainte_famille_amount
        temp_st_fam.border = border_default
        temp_st_fam.number_format = '#,##0.00$' 

    #total rev paroissiens
    total_par = sheet.cell(row=18, column=6)
    total_par.value = "TOTAL REVENUS DES PAROISSIENS"
    total_par.font = Font(bold=True,size=12)

    total_par_st_do = sheet.cell(row=18, column=9)
    total_par_st_do.value = '= SOMME(I7:I17)'

    total_par_st_fa = sheet.cell(row=18, column=10)
    total_par_st_fa.value = '= SOMME(J7:J17)'

    #AUTRES
    total_par = sheet.cell(row=20, column=6)
    total_par.value = "AUTRES"
    total_par.font = Font(bold=True,size=12)

    for x in range(11,16):
        tempAccount = sheet.cell(row=9+(x+1), column=6)
        tempAccount.value = list[x].account
        tempAccount.font = Font(bold=True)
        tempAccount.alignment = Alignment(horizontal="center")
        tempAccount.border = border_default

        temp_name = sheet.cell(row=9+(x+1), column=7)
        temp_name.value = list[x].name
        temp_name.border = border_default

        temp_st_do = sheet.cell(row=9+(x+1), column=9)
        temp_st_do.value = list[x].st_dominique_amount
        temp_st_do.border = border_default
        temp_st_do.number_format = '#,##0.00$' 

        temp_st_fam = sheet.cell(row=9+(x+1), column=10)
        temp_st_fam.value = list[x].sainte_famille_amount
        temp_st_fam.border = border_default
        temp_st_fam.number_format = '#,##0.00$' 

    #total rev autres
    total_autre = sheet.cell(row=26, column=6)
    total_autre.value = "TOTAL REVENUS DES AUTRES"
    total_autre.font = Font(bold=True,size=12)

    total_autre_st_do = sheet.cell(row=26, column=9)
    total_autre_st_do.value = '= SOMME(I21:I25)'

    total_autre_st_fa = sheet.cell(row=26, column=10)
    total_autre_st_fa.value = '= SOMME(J21:J25)'

    #total final revenus
    total_autre = sheet.cell(row=28, column=6)
    total_autre.value = "GRAND  TOTAL REVENUS "
    total_autre.font = Font(bold=True,size=16)

    total_autre_st_do = sheet.cell(row=28, column=9)
    total_autre_st_do.value = '= SOMME(I21:I25)'

    total_autre_st_fa = sheet.cell(row=28, column=10)
    total_autre_st_fa.value = '= SOMME(J21:J25)'


def WriteDepense(list, sheet):

     #revenus
    dep = sheet.cell(row=24, column=6)
    dep.value = "DÉPENSES"
    dep.font = Font(bold=True,size=14)

    for x in range(16,len(list)):
        tempAccount = sheet.cell(row=5+(x+4), column=6)
        tempAccount.value = list[x].account
        tempAccount.font = Font(bold=True)
        tempAccount.alignment = Alignment(horizontal="center")

        temp_name = sheet.cell(row=5+(x+4), column=7)
        temp_name.value = list[x].name

        temp_st_do = sheet.cell(row=5+(x+4), column=9)
        temp_st_do.value = list[x].st_dominique_amount

        temp_st_fa = sheet.cell(row=5+(x+4), column=10)
        temp_st_fa.value = list[x].sainte_famille_amount

def WriteExcel(list_operation):
    
    #open ExcelSheetWorker
    wb = openpyxl.Workbook()
    sheet = wb.active  

    bilan_lst = regroupement(list_operation)

    #set default col dimensions
    sheet.column_dimensions['G'].width = 40
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 20


    #titre en gras et en gros
    g2 = sheet.cell(row=2, column=7)
    g2.value = "REGROUPEMENT DES PAROISSES"
    g2.font = Font(bold=True,size=18)
    g2.alignment = Alignment(horizontal="center")

    #chiffre en haut des parroises
    for x in range(1,5):
        temp_cell = sheet.cell(row=3, column=x+8)
        temp_cell.value = x
        temp_cell. alignment = Alignment(horizontal="center")

    #date et parroisses
    d4 = sheet.cell(row=4, column=4)
    d4.value = "BUDGET année"
    d4.font = Font(bold=True,size=14)

    i4 = sheet.cell(row=4, column=9)
    i4.value = "SAINT-DOMINIQUE"
    i4.font = Font(bold=True)

    j4 = sheet.cell(row=4, column=10)
    j4.value = "SAINTE-FAMILLE"
    j4.font = Font(bold=True)


    #écriture des revenues
    WriteRevenus(bilan_lst, sheet)
    
    wb.save("demo.xlsx")


#regroupement
def regroupement(list_operations):
    grouped_list = [bilan_operations(40100,"QUÊTES",0,0),
                    bilan_operations(40200,"CAPITATION",0,0),
                    bilan_operations(40300,"LUMINAIRE (CULTE)",0,0),
                    bilan_operations(40400,"CÉLÉBRATIONS",0,0),
                    bilan_operations(40500,"QUÊTES COMMANDÉES",0,0),
                    bilan_operations(40600,"DONS",0,0),
                    bilan_operations(40700,"PASTORALE",0,0),
                    bilan_operations(40800,"OBJETS DE REVENTE(FEUILLET)",0,0),
                    bilan_operations(40900,"EXTRAIT DES ACTES",0,0),
                    bilan_operations(41000,"ACTICITÉ DE FINANCEMENT",0,0),
                    bilan_operations(49000,"AUTRES PROVENANT DES",0,0),
                    bilan_operations(50100,"LOCATIONS",0,0),
                    bilan_operations(50200,"INTÉRETS",0,0),
                    bilan_operations(50300,"SUBVENTION",0,0),
                    bilan_operations(50400,"ristournes assurance",0,0),
                    bilan_operations(59000,"revenus autres",0,0),
                    bilan_operations(60100,"SALAIRE ET C.E.SACRISTAIN",0,0),
                    bilan_operations(60200,"MINISTÈRE",0,0),
                    bilan_operations(60300,"FRAIS DE VOYAGE",0,0),
                    bilan_operations(60400,"CÉLÉBRATIONS",0,0),
                    bilan_operations(60500,"FEUILLET PAROISSIAL",0,0),
                    bilan_operations(60600,"cultes",0,0),
                    bilan_operations(60700,"UNITÉ DES DEUX-RIVES",0,0),
                    bilan_operations(60800,"ANIMATION LITURGIQUE",0,0),
                    bilan_operations(60900,"ANIMATION PASTORALE",0,0),
                    bilan_operations(61000,"PART ÉGLISE",0,0),
                    bilan_operations(61100,"OBJETS DE REVENTE",0,0),
                    bilan_operations(61200,"CIMETIÈRE",0,0),
                    bilan_operations(61300,"Quêtes commandéée",0,0),
                    bilan_operations(61400,"TRIBUT DIOCÉSAIN",0,0),
                    bilan_operations(70100,"SALAIRES C.E.",0,0),
                    bilan_operations(70200,"DÉPENSES DE BUREAU",0,0),
                    bilan_operations(70300,"HONORAIRES ET CONTRATS",0,0),
                    bilan_operations(70400,"FORMATION",0,0),
                    bilan_operations(70500,"ADMINISTRATION/TPS et TVQ",0,0),
                    bilan_operations(70600,"CIVILITÉES",0,0),
                    bilan_operations(79000,"AUTRE DÉPENSES DE BUREAU",0,0),
                    bilan_operations(80100,"SALAIRE ET C.E. EMPLOYEUR",0,0),
                    bilan_operations(80200,"CHAUFFAGE",0,0),
                    bilan_operations(80300,"ÉLECTRICITÉ",0,0),
                    bilan_operations(80400,"ENTRETIEN INTÉRIEUR",0,0),
                    bilan_operations(80500,"ENTRETIEN EXTÉRIEUR",0,0),
                    bilan_operations(80600,"RÉPARATIONS MAJEURES",0,0),
                    bilan_operations(80700,"ASSURANCE",0,0),
                    bilan_operations(89000,"AUTRE DÉPENSES SUR BÂTISSES",0,0)
                    ]
    
    for operation in list_operations:
        for bilan in grouped_list:
            if(int(str(operation.account)[:3]) == int(str(bilan.account)[:3])):
                if(operation.eglise == "SAINT-DOMINIQUE"):
                    bilan.st_dominique_amount = bilan.st_dominique_amount+operation.amount
                elif(operation.eglise == "SAINTE-FAMILLE"):
                    bilan.sainte_famille_amount = bilan.sainte_famille_amount+operation.amount

    return grouped_list



#prendre une capture écran de la premiere version avant de passer à la prochaine... pour éric
                
#debut programme
final_list = ReadAllExcel()
WriteExcel(final_list)