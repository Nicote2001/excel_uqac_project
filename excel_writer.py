import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from operation import Operations
from bilan_operation import bilan_operations
from bilan_total import bilan_totals

border_default=Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def write_paroissiens(list,sheet,total_bilan):
    '''
    Write the parroissiens sections.

            Parameters:
                    n (str): name of the file

            Returns:
                    binary_sum (str): Binary string of the sum of a and b
    '''
 #paroissiens
    f5 = sheet.cell(row=6, column=6)
    f5.value = "  PAROISSIENS"
    f5.font = Font(bold=True,size=12)

    for x in range(0,11):
        tempAccount = sheet.cell(row=6+(x+1), column=6)
        tempAccount.value = list[x].account
        tempAccount.font = Font(bold=True)
        tempAccount.alignment = Alignment(horizontal="center")
        tempAccount.border = border_default

        temp_name = sheet.cell(row=6+(x+1), column=7)
        temp_name.value = list[x].name
        temp_name.border = border_default

        temp_st_do = sheet.cell(row=6+(x+1), column=9)
        temp_st_do.value = list[x].st_dominique_amount
        temp_st_do.border = border_default
        temp_st_do.number_format = '#,##0.00$' 

        temp_st_fam = sheet.cell(row=6+(x+1), column=10)
        temp_st_fam.value = list[x].sainte_famille_amount
        temp_st_fam.border = border_default
        temp_st_fam.number_format = '#,##0.00$' 

        temp_st_ge = sheet.cell(row=6+(x+1), column=11)
        temp_st_ge.value = list[x].saint_gerard_amount
        temp_st_ge.border = border_default
        temp_st_ge.number_format = '#,##0.00$' 

        temp_st_th = sheet.cell(row=6+(x+1), column=12)
        temp_st_th.value = list[x].sainte_therese_amount
        temp_st_th.border = border_default
        temp_st_th.number_format = '#,##0.00$' 

    #total rev paroissiens
    total_par = sheet.cell(row=18, column=6)
    total_par.value = "TOTAL REVENUS DES PAROISSIENS"
    total_par.font = Font(bold=True,size=12)
    total_par.border = border_default

    total_par1 = sheet.cell(row=18, column=7)
    total_par1.border = border_default

    total_par_st_do = sheet.cell(row=18, column=9)
    total_par_st_do.value = total_bilan.paroissien_total.st_dominique
    total_par_st_do.number_format = '#,##0.00$' 
    total_par_st_do.font = Font(bold=True)
    total_par_st_do.border = border_default

    total_par_st_fa = sheet.cell(row=18, column=10)
    total_par_st_fa.value = total_bilan.paroissien_total.st_famille
    total_par_st_fa.number_format = '#,##0.00$' 
    total_par_st_fa.font = Font(bold=True)
    total_par_st_fa.border = border_default

    total_par_st_ge = sheet.cell(row=18, column=11)
    total_par_st_ge.value = total_bilan.paroissien_total.st_gerard
    total_par_st_ge.number_format = '#,##0.00$' 
    total_par_st_ge.font = Font(bold=True)
    total_par_st_ge.border = border_default

    total_par_st_th = sheet.cell(row=18, column=12)
    total_par_st_th.value = total_bilan.paroissien_total.st_therese
    total_par_st_th.number_format = '#,##0.00$' 
    total_par_st_th.font = Font(bold=True)
    total_par_st_th.border = border_default

def write_autre(list,sheet, total_bilan):
    total_par = sheet.cell(row=20, column=6)
    total_par.value = "AUTRES"
    total_par.font = Font(bold=True,size=12)
  
    for x in range(11,16):
        tempAccount = sheet.cell(row=9+(x+1), column=6)
        tempAccount.value = list[x].account
        tempAccount.font = Font(bold=True)
        tempAccount.alignment = Alignment(horizontal="center")
        tempAccount.border = border_default

        temp_name = sheet.cell(row=9+(x+1), column=7)
        temp_name.value = list[x].name
        temp_name.border = border_default

        temp_st_do = sheet.cell(row=9+(x+1), column=9)
        temp_st_do.value = list[x].st_dominique_amount
        temp_st_do.border = border_default
        temp_st_do.number_format = '#,##0.00$' 

        temp_st_fam = sheet.cell(row=9+(x+1), column=10)
        temp_st_fam.value = list[x].sainte_famille_amount
        temp_st_fam.border = border_default
        temp_st_fam.number_format = '#,##0.00$' 

        temp_st_ge = sheet.cell(row=9+(x+1), column=11)
        temp_st_ge.value = list[x].saint_gerard_amount
        temp_st_ge.border = border_default
        temp_st_ge.number_format = '#,##0.00$' 

        temp_st_th = sheet.cell(row=9+(x+1), column=12)
        temp_st_th.value = list[x].sainte_therese_amount
        temp_st_th.border = border_default
        temp_st_th.number_format = '#,##0.00$' 

    #total rev autres
    total_autre = sheet.cell(row=26, column=6)
    total_autre.value = "TOTAL REVENUS DES AUTRES"
    total_autre.font = Font(bold=True,size=12)
    total_autre.border = border_default

    total_par1 = sheet.cell(row=26, column=7)
    total_par1.border = border_default

    total_autre_st_do = sheet.cell(row=26, column=9)
    total_autre_st_do.value = total_bilan.autre_total.st_dominique
    total_autre_st_do.number_format = '#,##0.00$' 
    total_autre_st_do.font = Font(bold=True)
    total_autre_st_do.border = border_default

    total_autre_st_fa = sheet.cell(row=26, column=10)
    total_autre_st_fa.value = total_bilan.autre_total.st_famille
    total_autre_st_fa.number_format = '#,##0.00$' 
    total_autre_st_fa.font = Font(bold=True)
    total_autre_st_fa.border = border_default

    total_par_st_ge = sheet.cell(row=26, column=11)
    total_par_st_ge.value = total_bilan.autre_total.st_gerard
    total_par_st_ge.number_format = '#,##0.00$' 
    total_par_st_ge.font = Font(bold=True)
    total_par_st_ge.border = border_default

    total_par_st_th = sheet.cell(row=26, column=12)
    total_par_st_th.value = total_bilan.autre_total.st_therese
    total_par_st_th.number_format = '#,##0.00$' 
    total_par_st_th.font = Font(bold=True)
    total_par_st_th.border = border_default

def write_pastorale(list,sheet,total_bilan):
    #paroissiens
    f5 = sheet.cell(row=32, column=6)
    f5.value = "  PASTORALE"
    f5.font = Font(bold=True,size=12)

    for x in range(17,30):
        tempAccount = sheet.cell(row=15+(x+1), column=6)
        tempAccount.value = list[x].account
        tempAccount.font = Font(bold=True)
        tempAccount.alignment = Alignment(horizontal="center")
        tempAccount.border = border_default

        temp_name = sheet.cell(row=15+(x+1), column=7)
        temp_name.value = list[x].name
        temp_name.border = border_default

        temp_st_do = sheet.cell(row=15+(x+1), column=9)
        temp_st_do.value = list[x].st_dominique_amount
        temp_st_do.border = border_default
        temp_st_do.number_format = '#,##0.00$' 

        temp_st_fam = sheet.cell(row=15+(x+1), column=10)
        temp_st_fam.value = list[x].sainte_famille_amount
        temp_st_fam.border = border_default
        temp_st_fam.number_format = '#,##0.00$' 

        temp_st_ge = sheet.cell(row=15+(x+1), column=11)
        temp_st_ge.value = list[x].saint_gerard_amount
        temp_st_ge.border = border_default
        temp_st_ge.number_format = '#,##0.00$' 

        temp_st_th = sheet.cell(row=15+(x+1), column=12)
        temp_st_th.value = list[x].sainte_therese_amount
        temp_st_th.border = border_default
        temp_st_th.number_format = '#,##0.00$' 

    #total rev paroissiens
    total_par = sheet.cell(row=46, column=6)
    total_par.value = "TOTAL DÉPENSES DE PASTORALE"
    total_par.font = Font(bold=True,size=12)
    total_par.border = border_default

    total_par1 = sheet.cell(row=46, column=7)
    total_par1.border = border_default
    

    total_par_st_do = sheet.cell(row=46, column=9)
    total_par_st_do.value = total_bilan.pastorale_total.st_dominique
    total_par_st_do.number_format = '#,##0.00$' 
    total_par_st_do.font = Font(bold=True)
    total_par_st_do.border = border_default

    total_par_st_fa = sheet.cell(row=46, column=10)
    total_par_st_fa.value = total_bilan.pastorale_total.st_famille
    total_par_st_fa.number_format = '#,##0.00$' 
    total_par_st_fa.font = Font(bold=True)
    total_par_st_fa.border = border_default

    total_par_st_ge = sheet.cell(row=46, column=11)
    total_par_st_ge.value = total_bilan.pastorale_total.st_gerard
    total_par_st_ge.number_format = '#,##0.00$' 
    total_par_st_ge.font = Font(bold=True)
    total_par_st_ge.border = border_default

    total_par_st_th = sheet.cell(row=46, column=12)
    total_par_st_th.value = total_bilan.pastorale_total.st_therese
    total_par_st_th.number_format = '#,##0.00$' 
    total_par_st_th.font = Font(bold=True)
    total_par_st_th.border = border_default

def write_bureau(list,sheet, total_bilan):
    #AUTRES
    total_par = sheet.cell(row=48, column=6)
    total_par.value = " DE BUREAU"
    total_par.font = Font(bold=True,size=12)

    for x in range(30,37):
        
        tempAccount = sheet.cell(row=17+(x+1), column=6)
        tempAccount.value = list[x].account
        tempAccount.font = Font(bold=True)
        tempAccount.alignment = Alignment(horizontal="center")
        tempAccount.border = border_default

        temp_name = sheet.cell(row=17+(x+1), column=7)
        temp_name.value = list[x].name
        temp_name.border = border_default

        temp_st_do = sheet.cell(row=17+(x+1), column=9)
        temp_st_do.value = list[x].st_dominique_amount
        temp_st_do.border = border_default
        temp_st_do.number_format = '#,##0.00$' 

        temp_st_fam = sheet.cell(row=17+(x+1), column=10)
        temp_st_fam.value = list[x].sainte_famille_amount
        temp_st_fam.border = border_default
        temp_st_fam.number_format = '#,##0.00$' 

        temp_st_ge = sheet.cell(row=17+(x+1), column=11)
        temp_st_ge.value = list[x].saint_gerard_amount
        temp_st_ge.border = border_default
        temp_st_ge.number_format = '#,##0.00$' 

        temp_st_th = sheet.cell(row=17+(x+1), column=12)
        temp_st_th.value = list[x].sainte_therese_amount
        temp_st_th.border = border_default
        temp_st_th.number_format = '#,##0.00$' 

    #total rev autres
    total_autre = sheet.cell(row=55, column=6)
    total_autre.value = "TOTAL DÉPENSE DE BUREAU"
    total_autre.font = Font(bold=True,size=12)
    total_autre.border = border_default

    total_par1 = sheet.cell(row=55, column=7)
    total_par1.border = border_default

    total_autre_st_do = sheet.cell(row=55, column=9)
    total_autre_st_do.value = total_bilan.bureau_total.st_dominique
    total_autre_st_do.number_format = '#,##0.00$' 
    total_autre_st_do.font = Font(bold=True)
    total_autre_st_do.border = border_default

    total_autre_st_fa = sheet.cell(row=55, column=10)
    total_autre_st_fa.value = total_bilan.bureau_total.st_famille
    total_autre_st_fa.number_format = '#,##0.00$' 
    total_autre_st_fa.font = Font(bold=True)
    total_autre_st_fa.border = border_default

    total_par_st_ge = sheet.cell(row=55, column=11)
    total_par_st_ge.value = total_bilan.bureau_total.st_gerard
    total_par_st_ge.number_format = '#,##0.00$' 
    total_par_st_ge.font = Font(bold=True)
    total_par_st_ge.border = border_default

    total_par_st_th = sheet.cell(row=55, column=12)
    total_par_st_th.value = total_bilan.bureau_total.st_therese
    total_par_st_th.number_format = '#,##0.00$' 
    total_par_st_th.font = Font(bold=True)
    total_par_st_th.border = border_default

def WriteRevenus(list, sheet, total_final):

    #revenus
    f5 = sheet.cell(row=5, column=6)
    f5.value = "REVENUS"
    f5.font = Font(bold=True,size=14)

    #write parroisiens sections
    write_paroissiens(list,sheet, total_final)

    #write Autre section
    write_autre(list,sheet,total_final)
    
    #total final revenus
    total_autre = sheet.cell(row=28, column=6)
    total_autre.value = "GRAND  TOTAL REVENUS "
    total_autre.font = Font(bold=True,size=16)

    total_autre_st_do = sheet.cell(row=28, column=9)
    total_autre_st_do.value = total_final.revenus_total.st_dominique

    total_autre_st_fa = sheet.cell(row=28, column=10)
    total_autre_st_fa.value = total_final.revenus_total.st_famille

    total_autre_st_ge = sheet.cell(row=28, column=11)
    total_autre_st_ge.value = total_final.depenses_total.st_gerard

    total_autre_st_th = sheet.cell(row=28, column=12)
    total_autre_st_th.value = total_final.depenses_total.st_therese

def write_batisse(list, sheet, total_bilan):
    #de batisse
    total_par = sheet.cell(row=58, column=6)
    total_par.value = " DE BÂTISSE"
    total_par.font = Font(bold=True,size=12)

    for x in range(37,45):
        tempAccount = sheet.cell(row=20+(x+1), column=6)
        tempAccount.value = list[x].account
        tempAccount.font = Font(bold=True)
        tempAccount.alignment = Alignment(horizontal="center")
        tempAccount.border = border_default

        temp_name = sheet.cell(row=20+(x+1), column=7)
        temp_name.value = list[x].name
        temp_name.border = border_default

        temp_st_do = sheet.cell(row=20+(x+1), column=9)
        temp_st_do.value = list[x].st_dominique_amount
        temp_st_do.border = border_default
        temp_st_do.number_format = '#,##0.00$' 

        temp_st_fam = sheet.cell(row=20+(x+1), column=10)
        temp_st_fam.value = list[x].sainte_famille_amount
        temp_st_fam.border = border_default
        temp_st_fam.number_format = '#,##0.00$' 

        temp_st_ge = sheet.cell(row=20+(x+1), column=11)
        temp_st_ge.value = list[x].saint_gerard_amount
        temp_st_ge.border = border_default
        temp_st_ge.number_format = '#,##0.00$' 

        temp_st_th = sheet.cell(row=20+(x+1), column=12)
        temp_st_th.value = list[x].sainte_therese_amount
        temp_st_th.border = border_default
        temp_st_th.number_format = '#,##0.00$' 

        

    #total rev autres
    total_autre = sheet.cell(row=66, column=6)
    total_autre.value = "TOTAL DÉPENSE DE BÂTISSE"
    total_autre.font = Font(bold=True,size=12)
    total_autre.border = border_default

    total_par1 = sheet.cell(row=66, column=7)
    total_par1.border = border_default
    

    total_batisse_st_do = sheet.cell(row=66, column=9)
    total_batisse_st_do.value = total_bilan.batisse_total.st_dominique
    total_batisse_st_do.number_format = '#,##0.00$' 
    total_batisse_st_do.font = Font(bold=True)
    total_batisse_st_do.border = border_default
    

    total_batisse_st_fa = sheet.cell(row=66, column=10)
    total_batisse_st_fa.value = total_bilan.batisse_total.st_famille
    total_batisse_st_fa.number_format = '#,##0.00$' 
    total_batisse_st_fa.font = Font(bold=True)
    total_batisse_st_fa.border = border_default

    total_par_st_ge = sheet.cell(row=66, column=11)
    total_par_st_ge.value = total_bilan.batisse_total.st_gerard
    total_par_st_ge.number_format = '#,##0.00$' 
    total_par_st_ge.font = Font(bold=True)
    total_par_st_ge.border = border_default

    total_par_st_th = sheet.cell(row=66, column=12)
    total_par_st_th.value = total_bilan.batisse_total.st_therese
    total_par_st_th.number_format = '#,##0.00$' 
    total_par_st_th.font = Font(bold=True)
    total_par_st_th.border = border_default

def write_header(sheet):
    
    #titre en gras et en gros
    g2 = sheet.cell(row=2, column=7)
    g2.value = "REGROUPEMENT DES PAROISSES"
    g2.font = Font(bold=True,size=18)
    g2.alignment = Alignment(horizontal="center")

    #chiffre en haut des parroises
    for x in range(1,5):
        temp_cell = sheet.cell(row=3, column=x+8)
        temp_cell.value = x
        temp_cell. alignment = Alignment(horizontal="center")

    #date et parroisses
    d4 = sheet.cell(row=4, column=4)
    d4.value = "BUDGET année"
    d4.font = Font(bold=True,size=14)

    i4 = sheet.cell(row=4, column=9)
    i4.value = "SAINT-DOMINIQUE"
    i4.font = Font(bold=True)

    j4 = sheet.cell(row=4, column=10)
    j4.value = "SAINTE-FAMILLE"
    j4.font = Font(bold=True)

    k4 = sheet.cell(row=4, column=11)
    k4.value = "SAINT-GÉRARD"
    k4.font = Font(bold=True)

    k4 = sheet.cell(row=4, column=12)
    k4.value = "SAINTE-THÉRÈSE"
    k4.font = Font(bold=True)



def WriteDepense(list, sheet, list_bilan):

    #revenus
    f5 = sheet.cell(row=31, column=6)
    f5.value = "DÉPENSE"
    f5.font = Font(bold=True,size=14)

    #write Pastoral section
    write_pastorale(list,sheet,list_bilan)

    #write Bureau section
    write_bureau(list,sheet,list_bilan)
    
    #write Batisse section
    write_batisse(list,sheet,list_bilan)

    #total final revenus
    total_autre = sheet.cell(row=68, column=6)
    total_autre.value = "GRAND  TOTAL DÉPENSES "
    total_autre.font = Font(bold=True,size=16)

    total_autre_st_do = sheet.cell(row=68, column=9)
    total_autre_st_do.value = list_bilan.depenses_total.st_dominique

    total_autre_st_fa = sheet.cell(row=68, column=10)
    total_autre_st_fa.value = list_bilan.depenses_total.st_famille

    total_autre_st_ge = sheet.cell(row=68, column=11)
    total_autre_st_ge.value = list_bilan.depenses_total.st_gerard

    total_autre_st_th = sheet.cell(row=68, column=12)
    total_autre_st_th.value = list_bilan.depenses_total.st_therese

def write_footer(sheet,list_bilan):
    k73 = sheet.cell(row=73, column=6)
    k73.value = "GRAND  TOTAL REVENUS "
    k73.font = Font(bold=True,size=16)

    l73 = sheet.cell(row=73, column=9)
    l73.value = list_bilan.revenus_total.st_dominique
    l73.number_format = '#,##0.00$' 
    l73.font = Font(bold=True)
    l73.border = border_default

    i73 = sheet.cell(row=73, column=10)
    i73.value = list_bilan.revenus_total.st_famille
    i73.number_format = '#,##0.00$' 
    i73.font = Font(bold=True)
    i73.border = border_default
    
    i73b = sheet.cell(row=73, column=11)
    i73b.value = list_bilan.revenus_total.st_gerard
    i73b.number_format = '#,##0.00$' 
    i73b.font = Font(bold=True)
    i73b.border = border_default

    i73e = sheet.cell(row=73, column=12)
    i73e.value = list_bilan.revenus_total.st_therese
    i73e.number_format = '#,##0.00$' 
    i73e.font = Font(bold=True)
    i73e.border = border_default

    #depense
    j73 = sheet.cell(row=75, column=6)
    j73.value = "GRAND  TOTAL DÉPENSES "
    j73.font = Font(bold=True,size=16)

    m73 = sheet.cell(row=75, column=9)
    m73.value = list_bilan.depenses_total.st_dominique
    m73.number_format = '#,##0.00$' 
    m73.font = Font(bold=True)
    m73.border = border_default

    n73 = sheet.cell(row=75, column=10)
    n73.value = list_bilan.depenses_total.st_famille
    n73.number_format = '#,##0.00$' 
    n73.font = Font(bold=True)
    n73.border = border_default

    i73d = sheet.cell(row=75, column=11)
    i73d.value = list_bilan.depenses_total.st_gerard
    i73d.number_format = '#,##0.00$' 
    i73d.font = Font(bold=True)
    i73d.border = border_default


    i73c = sheet.cell(row=75, column=12)
    i73c.value = list_bilan.depenses_total.st_therese
    i73c.number_format = '#,##0.00$' 
    i73c.font = Font(bold=True)
    i73c.border = border_default

    #surplus annee
    o73 = sheet.cell(row=77, column=6)
    o73.value = "SURPLUS (DÉFICIT) DE L'ANNÉE PROJETÉ"
    o73.font = Font(bold=True,size=16)

    p73 = sheet.cell(row=77, column=9)
    p73.value = list_bilan.surplus_annee.st_dominique
    p73.number_format = '#,##0.00$' 
    p73.font = Font(bold=True)
    p73.border = border_default

    q73 = sheet.cell(row=77, column=10)
    q73.value = list_bilan.surplus_annee.st_famille
    q73.number_format = '#,##0.00$' 
    q73.font = Font(bold=True)
    q73.border = border_default

    t73c = sheet.cell(row=77, column=11)
    t73c.value = list_bilan.surplus_annee.st_gerard
    t73c.number_format = '#,##0.00$' 
    t73c.font = Font(bold=True)
    t73c.border = border_default

    t73b2 = sheet.cell(row=77, column=12)
    t73b2.value = list_bilan.surplus_annee.st_therese
    t73b2.number_format = '#,##0.00$' 
    t73b2.font = Font(bold=True)
    t73b2.border = border_default


    #surplus annee
    r73 = sheet.cell(row=79, column=6)
    r73.value = "SURPLUS (DÉFICIT) ACCUMULÉ AU 31/12/202#"
    r73.font = Font(bold=True,size=16)

    s73 = sheet.cell(row=79, column=9)
    s73.value = list_bilan.surplus_accumule.st_dominique

    t73 = sheet.cell(row=79, column=10)
    t73.value = list_bilan.surplus_accumule.st_famille

    t73b = sheet.cell(row=79, column=11)
    t73b.value = list_bilan.surplus_accumule.st_gerard

    t73b1 = sheet.cell(row=79, column=12)
    t73b1.value = list_bilan.surplus_accumule.st_therese

    #eglises de parroisses
    u73 = sheet.cell(row=79, column=6)
    u73.value = "Églises de la paroisse"
    u73.font = Font(bold=True,size=16)

    v73 = sheet.cell(row=79, column=9)
    v73.value = "SAINT-DOMINIQUE"

    w73 = sheet.cell(row=79, column=10)
    w73.value = "SAINTE-FAMILLE"

    x73 = sheet.cell(row=80, column=10)
    x73.value = "SAINT-CYRIAC"

    x73 = sheet.cell(row=81, column=10)
    x73.value = "SAINT-RAPHAEL"
    
    z73 = sheet.cell(row=79, column=11)
    z73.value = "SAINT-GÉRARD"

    a74 = sheet.cell(row=79, column=12)
    a74.value = "SAINTE-THÉRÈSE"

    b74 = sheet.cell(row=80, column=12)
    b74.value = "SAINT-MATHIAS"


def WriteExcel(bilan_lst):
    
    #open ExcelSheetWorker
    wb = openpyxl.Workbook()
    sheet = wb.active  

    #set default col dimensions
    sheet.column_dimensions['G'].width = 40
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 20

    #setup bilan_total_object
    total_final = bilan_totals(bilan_lst)
    total_final.set_totaux()

    #write header
    write_header(sheet)


    #écriture des revenues
    WriteRevenus(bilan_lst, sheet, total_final)
    
    WriteDepense(bilan_lst,sheet, total_final)

    #write the footer
    write_footer(sheet,total_final)
    wb.save("demo.xlsx")

#regroupement
def regroupement(list_operations):

    try:
        grouped_list = [bilan_operations(4010000,"QUÊTES ",1,0,0),
                        bilan_operations(4020000,"CAPITATION",1,0,0),
                        bilan_operations(4030000,"TRIBUT",1,0,0),
                        bilan_operations(4040000,"LUMINAIRE",1,0,0),
                        bilan_operations(4050000,"CÉLÉBRATIONS",1,0,0),
                        bilan_operations(4060000,"QUÊTES COMMANDÉES",1,0,0),
                        bilan_operations(4070000,"DONS",1,0,0),
                        bilan_operations(4080000,"PASTORALE",1,0,0),
                        bilan_operations(4090000,"OBJETS DE REVENTE",1,0,0),
                        bilan_operations(4100000,"EXTRAITS DE NAISSANCE",1,0,0),
                        bilan_operations(4110000,"ACTIVITÉS DE FINANCEMENT",1,0,0),
                        bilan_operations(5010000,"LOCATIONS",1,0,0),
                        bilan_operations(5020000,"INTÉRETS",1,0,0),
                        bilan_operations(5030000,"SUBVENTION",1,0,0),
                        bilan_operations(5990000,"REVENUS AUTRES",1,0,0),
                        bilan_operations(6010000,"SALAIRES",0,0,0),
                        bilan_operations(6020000,"MINISTÈRE, PRÊTRE ET DIACRE",0,0,0),
                        bilan_operations(6030000,"FRAIS DE VOYAGE",0,0,0),
                        bilan_operations(6040000,"CÉLÉBRATIONS",0,0,0),
                        bilan_operations(6050000,"FEUILLET PAROISSIAL",0,0,0),
                        bilan_operations(6060000,"CULTES",0,0,0),
                        bilan_operations(6070000,"UNITÉ DES DEUX-RIVES",0,0,0),
                        bilan_operations(6080000,"ANIMATION LITURGIQUE ET PASTORALE ET CHANTS",1,0,0),
                        bilan_operations(6090000,"PART ÉGLISE",0,0,0),
                        bilan_operations(6100000,"OBJETS DE REVENTE",0,0,0),
                        bilan_operations(6110000,"CIMETIÈRE",0,0,0),
                        bilan_operations(6120000,"Quêtes commandéée",0,0,0),
                        bilan_operations(6130000,"TRIBUT DIOCÉSAIN",0,0,0),
                         bilan_operations(6990000,"AUTRES DÉPENSES PASTORALE",0,0,0),
                        bilan_operations(7010000,"SALAIRES C.E.",0,0,0),
                        bilan_operations(7020000,"DÉPENSES DE BUREAU",0,0,0),
                        bilan_operations(7030000,"HONORAIRES ET CONTRATS",0,0,0),
                        bilan_operations(7040000,"FORMATION",0,0,0),
                        bilan_operations(7050000,"FRAIS DE BANQUE",0,0,0),
                        bilan_operations(7060000,"TPS et TVQ",0,0,0),
                        bilan_operations(7070000,"CIVILITÉES",0,0,0),
                        bilan_operations(7990000,"AUTRE DÉPENSES DE BUREAU",0,0,0),
                        bilan_operations(8010000,"SALAIRE ET C.E. EMPLOYEUR",0,0,0),
                        bilan_operations(8020000,"CHAUFFAGE",0,0,0),
                        bilan_operations(8030000,"ÉLECTRICITÉ",0,0,0),
                        bilan_operations(8040000,"ENTRETIEN INTÉRIEUR",0,0,0),
                        bilan_operations(8050000,"ENTRETIEN EXTÉRIEUR",0,0,0),
                        bilan_operations(8060000,"RÉPARATIONS MINEURES",0,0,0),
                        bilan_operations(8070000,"ASSURANCE",0,0,0),
                        bilan_operations(8990000,"AUTRE DÉPENSES SUR BÂTISSES",0,0,0),
                        bilan_operations(9010000,"IMMOBILISATION",0,0,0)
                        ]
        
        for operation in list_operations:
            for bilan in grouped_list:
                if(int(str(operation.account)[:3]) == int(str(bilan.account)[:3])):
                    if operation.type != bilan.type:
                        operation.amount = 0 - operation.amount

                    if(operation.eglise == "SAINT-DOMINIQUE"):
                        bilan.st_dominique_amount = bilan.st_dominique_amount+operation.amount
                    elif(operation.eglise == "SAINTE-FAMILLE"):
                        bilan.sainte_famille_amount = bilan.sainte_famille_amount+operation.amount
                    elif(operation.eglise == "SAINT-GERARD"):
                        bilan.saint_gerard_amount = bilan.saint_gerard_amount+operation.amount
                    elif(operation.eglise == "SAINTE-THERESE"):
                        bilan.sainte_therese_amount = bilan.sainte_therese_amount+operation.amount

        return grouped_list
    except:
        print("une erreur lors de la compilation des résultat groupée s'est produite")
