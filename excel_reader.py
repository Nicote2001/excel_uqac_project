import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from operation import Operations
from bilan_operation import bilan_operations

def ExcelfileToList(path,nom_eglise):
    '''
    Returns list from a excel ( formated ).

            Parameters:
                    n (str): name of the file

            Returns:
                    binary_sum (str): Binary string of the sum of a and b
    '''
    file = path
    dataframe = openpyxl.load_workbook(file)
    
    # Define variables
    dataframe1 = dataframe.active
    list = []
    temp_amount=0
    temp_name=""
    temp_no_account =0
    temp_type=0 # 0 = revenue, 1= depense
    
    # Iteration dans les row
    for row1 in range(2, dataframe1.max_row):
            if dataframe1.cell(row=row1,column = 9).value  != None and isinstance(dataframe1.cell(row=row1,column = 9).value, (int, float)):
                temp_no_account = dataframe1.cell(row = row1,column = 1).value
                temp_name = dataframe1.cell(row = row1,column = 2).value
                temp_amount = dataframe1.cell(row = row1,column = 9).value
                temp_type = 0
                list.append(Operations(temp_no_account,temp_name,temp_type,temp_amount,nom_eglise)) #ajout a la liste
            elif dataframe1.cell(row=row1,column = 10).value  != None and isinstance(dataframe1.cell(row=row1,column = 10).value, (int, float)):
                temp_no_account = dataframe1.cell(row = row1,column = 1).value
                temp_name = dataframe1.cell(row = row1,column = 2).value
                temp_amount = dataframe1.cell(row = row1,column = 10).value
                temp_type = 1
                list.append(Operations(temp_no_account,temp_name,temp_type,temp_amount,nom_eglise)) #ajout a la liste
                        
            
    return list

def ReadAllExcel(list):
    '''
    Read all excels, made and group list and return final list

            Parameters:

            Returns:
                    final_list (List[operation])
    '''
    try:
        st_do_lst  = ExcelfileToList(list[0], "SAINT-DOMINIQUE")
        st_fa_lst  = ExcelfileToList(list[1],'SAINTE-FAMILLE')
        st_ge_lst  = ExcelfileToList(list[2], "SAINT-GERARD")
        st_th_lst  = ExcelfileToList(list[3], "SAINTE-THERESE")
        final_lst = st_do_lst + st_fa_lst + st_ge_lst + st_th_lst

        for x in final_lst:
            print(x.name+" - montant : "+str(x.amount))
        return final_lst
    except:
        print("il y a eu une erreur lors de la lecture d'un des fichiers")